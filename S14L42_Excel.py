import os
import openpyxl

os.chdir(r'C:\Users\benne\OneDrive\Documents\GitHub\Automate_the_boring_stuff')

# Open an existing workbook
workbook = openpyxl.load_workbook('example.xlsx')
# Print names of worksheets
print(workbook.sheetnames)
# Access sheet by name
sheet = workbook['Sheet1']
# Access cell value
cell = sheet['A1']
# Print values in various ways from cell object
print(cell.value)
print(str(cell.value))
print(sheet['B1'].value)
print(sheet.cell(row=1, column=2).value)

# Use for loop to print cell values
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

# Open a new empty workbook in memory
wb = openpyxl.Workbook()
# Print default sheet name of new workbook
print(wb.sheetnames)
# Access sheet object
sheet = wb['Sheet']
# Assign values to cells in sheet object
sheet['A1'] = 42
sheet['A2'] = 'Hello'
# Save workbook object as an excel workbook file
wb.save('example_2.xlsx')
# Create new sheet
sheet2 = wb.create_sheet()
# Change sheet default name
sheet2.title = 'My New Sheet'
wb.save('example_3.xlsx')
# Create sheet in first position and name it
wb.create_sheet(index=0, title='My other sheet')
wb.save('example_4.xlsx')
